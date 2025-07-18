# 依赖项安装指南：
# 请确保已安装以下库：
# pip install streamlit numpy scipy matplotlib pandas openpyxl matplotlib-font-manager

import streamlit as st
import numpy as np
from scipy.integrate import odeint
import matplotlib.pyplot as plt
import pandas as pd
import matplotlib.font_manager as fm
from io import BytesIO
import matplotlib as mpl
import os
import urllib.request

# 设置全局字体以支持中文
try:
    # 获取当前文件所在目录
    current_dir = os.path.dirname(os.path.abspath(__file__))
    fonts_dir = os.path.join(current_dir, 'fonts')

    # 确保 fonts 目录存在
    if not os.path.exists(fonts_dir):
        os.makedirs(fonts_dir)

    # 检查字体文件是否存在，如果不存在则尝试下载
    simhei_path = os.path.join(fonts_dir, 'simhei.ttf')
    msyh_path = os.path.join(fonts_dir, 'msyh.ttf')

    if not os.path.exists(simhei_path):
        # 从 GitHub 下载 SimHei 替代字体
        simhei_url = "https://github.com/googlefonts/noto-cjk/raw/main/Sans/OTF/SimplifiedChinese/NotoSansCJKsc-Regular.otf"
        urllib.request.urlretrieve(simhei_url, simhei_path)

    if not os.path.exists(msyh_path):
        # 从 GitHub 下载 Microsoft YaHei 替代字体
        msyh_url = "https://github.com/googlefonts/noto-cjk/raw/main/Sans/OTF/SimplifiedChinese/NotoSansCJKsc-Regular.otf"
        urllib.request.urlretrieve(msyh_url, msyh_path)

    # 添加字体目录到字体路径
    font_files = fm.findSystemFonts(fontpaths=[fonts_dir])
    for font_file in font_files:
        fm.fontManager.addfont(font_file)

    # 设置中文字体
    plt.rcParams['font.sans-serif'] = ['Noto Sans SC', 'SimHei', 'Microsoft YaHei', 'Arial Unicode MS', 'sans-serif']
    plt.rcParams['axes.unicode_minus'] = False

    # 验证字体是否加载成功
    zh_font = fm.FontProperties(fname=simhei_path)

    # 显示成功消息
    st.sidebar.success("中文字体已成功加载")

except Exception as e:
    # 如果找不到中文字体，使用默认字体
    st.sidebar.warning(f"无法加载中文字体: {str(e)}，图表中文显示可能异常")
    zh_font = fm.FontProperties()
    plt.rcParams['font.sans-serif'] = ['DejaVu Sans', 'Arial Unicode MS', 'sans-serif']
    plt.rcParams['axes.unicode_minus'] = False

# 检查 Streamlit 版本
try:
    import streamlit as st

    if st.__version__ < '1.0.0':
        st.warning("请升级 Streamlit 到最新版本以获得最佳体验。")
except ImportError:
    st.error("Streamlit 未安装，请使用 'pip install streamlit' 安装。")

st.set_page_config(page_title="乳糖水解动力学模拟 - 教学版", layout="wide")

# 语言选择
language = st.sidebar.selectbox("选择语言 / Select Language", ["中文", "English"])
lang = "zh" if language == "中文" else "en"

# 翻译字典
translations = {
    "zh": {
        "title": "🍼 乳糖水解动力学模拟 - 教学版",
        "intro": """
        ### 欢迎体验乳糖水解模拟
        乳糖水解是乳糖在β-半乳糖苷酶作用下分解为半乳糖和葡萄糖的过程，广泛应用于食品工业（如乳糖不耐受产品的生产）。本工具通过动力学模型模拟这一过程，帮助你理解酶催化反应和产物抑制的影响。

        **学习目标：**
        - 掌握Michaelis-Menten动力学的基本原理。
        - 理解产物抑制（半乳糖抑制）如何影响反应速率。
        - 通过交互式模拟，探索参数对乳糖水解的影响。
        """,
        "model_desc": "该模型模拟乳糖在β-半乳糖苷酶作用下的水解过程，考虑产物抑制效应（半乳糖抑制）。",
        "equation": r"""
        **三种抑制类型的动力学方程：**

        **1. 竞争性抑制：** 
        $$ r = \frac{V_{max} \cdot L}{K_m \cdot (1 + \frac{Gal}{K_i}) + L} $$

        **2. 非竞争性抑制：** 
        $$ r = \frac{V_{max} \cdot L}{(K_m + L) \cdot (1 + \frac{Gal}{K_i})} $$

        **3. 反竞争性抑制：** 
        $$ r = \frac{V_{max} \cdot L}{K_m + L \cdot (1 + \frac{Gal}{K_i})} $$
        """,
        "gal_desc": "其中 $Gal = L_0 - L$ 表示生成的半乳糖浓度",
        "reaction_params": "反应参数",
        "initial_lactose": "初始乳糖浓度 (mM) - 反应开始时的乳糖量",
        "enzyme_conc": "酶浓度 (U/mL) - 决定反应速率的关键因素",
        "reaction_time": "反应时间 (小时) - 模拟的总时间",
        "kinetic_params": "动力学参数",
        "km": "Km (mM) - 米氏常数，表示酶对底物的亲和力",
        "ki": "Ki (mM) - 抑制常数，表示半乳糖的抑制强度",
        "steps": "模拟精度 - 数值计算的步数",
        "theory": "理论背景",
        "theory_content": """
        **动力学模型：**
        - $L$: 乳糖浓度 (mM)
        - $Gal$: 半乳糖浓度 (mM)
        - $V_{max}$: 最大反应速率，与酶浓度 ($E$) 成正比
        - $K_m$: 米氏常数，表示酶对底物的亲和力（$K_m$ 越小，亲和力越高）
        - $K_i$: 产物抑制常数，表示半乳糖对酶的抑制强度（$K_i$ 越小，抑制越强）
        """,
        "equation_desc": "上述方程考虑了产物半乳糖对酶活的不同抑制机制。",
        "inhibition_type": "产物抑制类型",
        "competitive": "竞争性抑制",
        "non_competitive": "非竞争性抑制",
        "uncompetitive": "反竞争性抑制",
        "inhibition_types_desc": {
            "competitive": "抑制剂与底物竞争酶的活性位点",
            "non_competitive": "抑制剂结合在酶的其他部位，降低酶活性",
            "uncompetitive": "抑制剂只与酶-底物复合物结合"
        },
        "compare_inhibition": "比较有无产物抑制的模拟结果",
        "final_lactose": "最终乳糖浓度",
        "final_galactose": "最终半乳糖浓度",
        "conversion_rate": "转化率",
        "download_data": "下载模拟数据 (Excel)",
        "rate_analysis": "反应速率分析",
        "max_rate": "最大反应速率: **{:.2f} mM/小时** (发生在 {:.1f} 小时)",
        "exercises": "练习题",
        "exercise_content": """
        1. 如果酶浓度加倍，乳糖水解速率会如何变化？尝试调整参数并观察结果。
        2. 在什么条件下，产物抑制对反应速率的影响最小？调整 $K_i$ 值并分析。
        3. 使用模拟工具，找到使转化率达到90%所需的最短反应时间。
        """,
        "error": "计算错误: {}",
        "copyright": "© 生物反应工程教学模拟器 | 基于Michaelis-Menten动力学与产物抑制模型",
        "lb_chart": "Lineweaver-Burk 图表",
        "fixed_galactose": "固定半乳糖浓度 (mM)",
        "lb_explanation": {
            "competitive": "蓝色线条表示无抑制剂情况，遵循标准Michaelis-Menten动力学。红色线条表示固定半乳糖浓度下的竞争性抑制。注意两条线在y轴上的交点相同（绿色点），这表明竞争性抑制不影响 $V_{{max}}$，但改变了表观 $K_m$（与X轴负半轴的交点不同，蓝色和红色星号）。",
            "non_competitive": "蓝色线条表示无抑制剂情况，遵循标准Michaelis-Menten动力学。红色线条表示固定半乳糖浓度下的非竞争性抑制。注意两条线在x轴上的交点相同（绿色星号），这表明非竞争性抑制不影响 $K_m$，但改变了表观 $V_{{max}}$（与y轴的交点不同）。",
            "uncompetitive": "蓝色线条表示无抑制剂情况，遵循标准Michaelis-Menten动力学。红色线条表示固定半乳糖浓度下的反竞争性抑制。注意两条线平行（斜率相同），这表明反竞争性抑制同时改变了 $K_m$ 和 $V_{{max}}$，但斜率不变。"
        },
        "time_label": "时间 (小时)",
        "concentration_label": "浓度 (mM)",
        "substrate_label": "底物浓度 L (mM)",
        "rate_label": "反应速率 (mM/小时)",
        "no_inhibition": "无抑制"
    },
    "en": {
        "title": "🍼 Lactose Hydrolysis Kinetics Simulation - Educational Version",
        "intro": """
        ### Welcome to Lactose Hydrolysis Simulation
        Lactose hydrolysis is the process where lactose is broken down into galactose and glucose by β-galactosidase, widely used in the food industry (e.g., lactose-free products). This tool simulates this process using a kinetic model, helping you understand enzyme catalysis and product inhibition effects.

        **Learning Objectives:**
        - Understand the basics of Michaelis-Menten kinetics.
        - Explore how product inhibition (galactose) affects reaction rates.
        - Investigate parameter effects on lactose hydrolysis through interactive simulation.
        """,
        "model_desc": "This model simulates lactose hydrolysis by β-galactosidase, considering product inhibition (galactose inhibition).",
        "equation": r"""
        **Kinetic Equations for Three Inhibition Types:**

        **1. Competitive Inhibition:** 
        $$ r = \frac{V_{max} \cdot L}{K_m \cdot (1 + \frac{Gal}{K_i}) + L} $$

        **2. Non-competitive Inhibition:** 
        $$ r = \frac{V_{max} \cdot L}{(K_m + L) \cdot (1 + \frac{Gal}{K_i})} $$

        **3. Uncompetitive Inhibition:** 
        $$ r = \frac{V_{max} \cdot L}{K_m + L \cdot (1 + \frac{Gal}{K_i})} $$
        """,
        "gal_desc": "where $Gal = L_0 - L$ represents the concentration of produced galactose",
        "reaction_params": "Reaction Parameters",
        "initial_lactose": "Initial Lactose Concentration (mM) - Amount of lactose at the start",
        "enzyme_conc": "Enzyme Concentration (U/mL) - Key factor determining reaction rate",
        "reaction_time": "Reaction Time (hours) - Total simulation duration",
        "kinetic_params": "Kinetic Parameters",
        "km": "Km (mM) - Michaelis Constant, indicating enzyme-substrate affinity",
        "ki": "Ki (mM) - Inhibition Constant, indicating galactose inhibition strength",
        "steps": "Simulation Precision - Number of calculation points",
        "theory": "Theoretical Background",
        "theory_content": """
        **Kinetic Model:**
        - $L$: Lactose concentration (mM)
        - $Gal$: Galactose concentration (mM)
        - $V_{max}$: Maximum reaction rate, proportional to enzyme concentration ($E$)
        - $K_m$: Michaelis constant, indicating enzyme-substrate affinity (lower $K_m$, higher affinity)
        - $K_i$: Product inhibition constant, indicating galactose inhibition strength (lower $K_i$, stronger inhibition)
        """,
        "equation_desc": "These equations account for different inhibition mechanisms of the enzyme by the product galactose.",
        "inhibition_type": "Product Inhibition Type",
        "competitive": "Competitive Inhibition",
        "non_competitive": "Non-competitive Inhibition",
        "uncompetitive": "Uncompetitive Inhibition",
        "inhibition_types_desc": {
            "competitive": "Inhibitor competes with substrate for active site",
            "non_competitive": "Inhibitor binds to enzyme at different site, reducing activity",
            "uncompetitive": "Inhibitor binds only to enzyme-substrate complex"
        },
        "compare_inhibition": "Compare Simulation with and without Product Inhibition",
        "final_lactose": "Final Lactose Concentration",
        "final_galactose": "Final Galactose Concentration",
        "conversion_rate": "Conversion Rate",
        "download_data": "Download Simulation Data (Excel)",
        "rate_analysis": "Reaction Rate Analysis",
        "max_rate": "Maximum Reaction Rate: **{:.2f} mM/hour** (occurs at {:.1f} hours)",
        "exercises": "Exercises",
        "exercise_content": """
        1. How does doubling the enzyme concentration affect the hydrolysis rate? Adjust the parameters and observe.
        2. Under what conditions is the effect of product inhibition minimal? Adjust $K_i$ and analyze.
        3. Use the tool to find the shortest reaction time needed for a 90% conversion rate.
        """,
        "error": "Calculation Error: {}",
        "copyright": "© Bioreaction Engineering Educational Simulator | Based on Michaelis-Menten Kinetics with Product Inhibition Model",
        "lb_chart": "Lineweaver-Burk Plot",
        "fixed_galactose": "Fixed Galactose Concentration (mM)",
        "lb_explanation": {
            "competitive": "Blue line represents no inhibitor case, following standard Michaelis-Menten kinetics. Red line represents competitive inhibition at fixed galactose concentration. Note that both lines intersect at the same point on the y-axis (green point), indicating that competitive inhibition does not affect $V_{{max}}$, but changes the apparent $K_m$ (different intercepts on the negative x-axis, blue and red stars).",
            "non_competitive": "Blue line represents no inhibitor case, following standard Michaelis-Menten kinetics. Red line represents non-competitive inhibition at fixed galactose concentration. Note that both lines intersect at the same point on the x-axis (green star), indicating that non-competitive inhibition does not affect $K_m$, but changes the apparent $V_{{max}}$ (different intercepts on the y-axis).",
            "uncompetitive": "Blue line represents no inhibitor case, following standard Michaelis-Menten kinetics. Red line represents uncompetitive inhibition at fixed galactose concentration. Note that both lines are parallel (same slope), indicating that uncompetitive inhibition changes both $K_m$ and $V_{{max}}$, but the slope remains constant."
        },
        "time_label": "Time (hours)",
        "concentration_label": "Concentration (mM)",
        "substrate_label": "Substrate Concentration L (mM)",
        "rate_label": "Reaction Rate (mM/hour)",
        "no_inhibition": "No Inhibition"
    }
}

t = translations[lang]

st.title(t["title"])
st.markdown(t["intro"])

# 在侧边栏添加抑制类型选择
inhibition_types = st.sidebar.multiselect(
    t["inhibition_type"],
    options=[t["competitive"], t["non_competitive"], t["uncompetitive"]],
    default=[t["competitive"], t["non_competitive"], t["uncompetitive"]],
    help="选择要模拟的抑制类型"
)

# 添加抑制类型描述
if lang == "zh":
    st.sidebar.markdown("**抑制类型说明:**")
    st.sidebar.markdown(f"- **{t['competitive']}**: {t['inhibition_types_desc']['competitive']}")
    st.sidebar.markdown(f"- **{t['non_competitive']}**: {t['inhibition_types_desc']['non_competitive']}")
    st.sidebar.markdown(f"- **{t['uncompetitive']}**: {t['inhibition_types_desc']['uncompetitive']}")
else:
    st.sidebar.markdown("**Inhibition Type Descriptions:**")
    st.sidebar.markdown(f"- **{t['competitive']}**: {t['inhibition_types_desc']['competitive']}")
    st.sidebar.markdown(f"- **{t['non_competitive']}**: {t['inhibition_types_desc']['non_competitive']}")
    st.sidebar.markdown(f"- **{t['uncompetitive']}**: {t['inhibition_types_desc']['uncompetitive']}")

# 主界面
st.markdown(t["model_desc"])

# 添加米氏方程介绍
st.markdown("### 米氏方程 (Michaelis-Menten Equation)" if lang == "zh" else "### Michaelis-Menten Equation")
st.markdown(r"酶催化反应的基本动力学方程：" if lang == "zh" else "The fundamental kinetic equation for enzyme-catalyzed reactions:")
st.markdown(r"$$ r = \frac{V_{max} \cdot L}{K_m + L} $$")
st.markdown(r"其中：" if lang == "zh" else "Where:")
st.markdown(r"- $r$: 反应速率 (mM/小时)" if lang == "zh" else "- $r$: Reaction rate (mM/hour)")
st.markdown(r"- $V_{max}$: 最大反应速率 (mM/小时)" if lang == "zh" else "- $V_{max}$: Maximum reaction rate (mM/hour)")
st.markdown(r"- $L$: 底物浓度 (mM)" if lang == "zh" else "- $L$: Substrate concentration (mM)")
st.markdown(r"- $K_m$: 米氏常数 (mM)，表示酶对底物的亲和力" if lang == "zh" else "- $K_m$: Michaelis constant (mM), indicating enzyme-substrate affinity")

# 添加产物抑制模型介绍
st.markdown("### 产物抑制模型" if lang == "zh" else "### Product Inhibition Model")
st.markdown(t["equation"])  # 显示三种抑制类型的方程
st.markdown(t["gal_desc"])

# 创建两列布局
col1, col2 = st.columns(2)

with col1:
    st.header(t["reaction_params"])
    L0 = st.slider(
        label=t["initial_lactose"],
        min_value=0.1,
        max_value=500.0,
        value=200.0,
        step=0.1,
        help="反应开始时的乳糖浓度 (mM)" if lang == "zh" else "Initial lactose concentration (mM)"
    )
    E = st.slider(
        label=t["enzyme_conc"],
        min_value=0.001,
        max_value=10.0,
        value=1.0,
        step=0.001,
        help="酶浓度 (U/mL)，影响反应速率" if lang == "zh" else "Enzyme concentration (U/mL), key factor determining reaction rate"
    )
    t_max = st.slider(
        label=t["reaction_time"],
        min_value=0.0,
        max_value=12.0,
        value=1.0,
        step=0.01,
        help="模拟的反应时间 (小时)" if lang == "zh" else "Simulation reaction time (hours)"
    )

with col2:
    st.header(t["kinetic_params"])
    Km = st.slider(
        label=t["km"],
        min_value=0.1,
        max_value=50.0,
        value=30.0,
        step=0.1,
        help="米氏常数 (mM)，表示酶对底物的亲和力" if lang == "zh" else "Michaelis constant (mM), indicating enzyme-substrate affinity"
    )
    Ki = st.slider(
        label=t["ki"],
        min_value=0.1,
        max_value=50.0,
        value=10.0,
        step=0.1,
        help="抑制常数 (mM)，表示半乳糖的抑制强度" if lang == "zh" else "Inhibition constant (mM), indicating galactose inhibition strength"
    )
    steps = st.slider(
        label=t["steps"],
        min_value=50,
        max_value=500,
        value=200,
        step=1,
        help="数值计算的步数，影响模拟精度" if lang == "zh" else "Number of calculation points, affects simulation precision"
    )

# 理论背景
with st.expander(t["theory"]):
    st.markdown(t["theory_content"])
    st.markdown(t["equation_desc"])


# 模拟函数 - 修改为支持多种抑制类型
@st.cache_data
def solve_model(L0, Vmax, Km, Ki, t_max, steps, inhibition_type):
    if L0 <= 0 or Vmax <= 0 or Km <= 0 or Ki <= 0:
        raise ValueError("参数必须为正数")
    t_min = np.linspace(0, t_max * 60, steps)
    sol = odeint(model, L0, t_min, args=(Vmax, Km, Ki, L0, inhibition_type))
    L = sol[:, 0]
    Gal = np.maximum(L0 - L, 0)
    t_hour = t_min / 60
    rates = np.abs(np.gradient(L, t_hour))
    return t_hour, L, Gal, rates


def model(L, t, Vmax, Km, Ki, L0, inhibition_type):
    L = max(L, 1e-6)
    Gal = L0 - L

    # 根据抑制类型选择不同的动力学方程
    if inhibition_type == "competitive":
        denominator = Km * (1 + Gal / Ki) + L
    elif inhibition_type == "non_competitive":
        denominator = (Km + L) * (1 + Gal / Ki)
    elif inhibition_type == "uncompetitive":
        denominator = Km + L * (1 + Gal / Ki)
    else:
        denominator = Km + L  # 无抑制

    dLdt = -Vmax * L / denominator
    return dLdt


try:
    Vmax = E
    # 创建颜色映射
    colors = {
        "competitive": '#4E6691',
        "non_competitive": '#4D8B31',
        "uncompetitive": '#B8474D',
        "no_inhibition": '#808080'
    }

    # 添加浓度-时间分析标题
    st.subheader("浓度-时间分析" if lang == "zh" else "Concentration-Time Profile")
    
    # 可视化
    fig, ax = plt.subplots(figsize=(10, 6))

    # 存储所有模拟结果
    all_results = {}

    # 处理无抑制情况
    t_hour_no_inh, L_no_inh, Gal_no_inh, rates_no_inh = solve_model(L0, Vmax, Km, Ki, t_max, steps, "no_inhibition")
    all_results["no_inhibition"] = (t_hour_no_inh, L_no_inh, Gal_no_inh, rates_no_inh)
    ax.plot(t_hour_no_inh, L_no_inh, '--', color=colors["no_inhibition"], linewidth=2.5,
            label=f"乳糖 ({t['no_inhibition']})" if lang == "zh" else f"Lactose ({t['no_inhibition']})")
    ax.plot(t_hour_no_inh, Gal_no_inh, '--', color='#FF7F0E', linewidth=2.5,
            label=f"半乳糖 ({t['no_inhibition']})" if lang == "zh" else f"Galactose ({t['no_inhibition']})")

    # 处理选中的抑制类型
    for itype in inhibition_types:
        # 将显示名称映射到内部标识符
        if itype == t["competitive"]:
            key = "competitive"
            label_prefix = "竞争性" if lang == "zh" else "Competitive"
        elif itype == t["non_competitive"]:
            key = "non_competitive"
            label_prefix = "非竞争性" if lang == "zh" else "Non-competitive"
        elif itype == t["uncompetitive"]:
            key = "uncompetitive"
            label_prefix = "反竞争性" if lang == "zh" else "Uncompetitive"
        else:
            continue

        t_hour, L, Gal, rates = solve_model(L0, Vmax, Km, Ki, t_max, steps, key)
        all_results[key] = (t_hour, L, Gal, rates)

        # 绘制乳糖和半乳糖曲线
        ax.plot(t_hour, L, color=colors[key], linewidth=2.5,
                label=f"乳糖 ({label_prefix}抑制)" if lang == "zh" else f"Lactose ({label_prefix} Inhibition)")
        ax.plot(t_hour, Gal, color=colors[key], linestyle=':', linewidth=2.5,
                label=f"半乳糖 ({label_prefix}抑制)" if lang == "zh" else f"Galactose ({label_prefix} Inhibition)")

        # 添加转化率标注
        conversion = (1 - L[-1] / L0) * 100
        ax.annotate(f'{conversion:.1f}% {t["conversion_rate"]}',
                    xy=(t_hour[-1], Gal[-1]),
                    xytext=(t_hour[-1] - 0.2, Gal[-1] + 0.05 * L0),
                    arrowprops=dict(arrowstyle='->', color=colors[key]),
                    fontsize=10, color=colors[key], fontproperties=zh_font if lang == "zh" else None)

    # 使用翻译的坐标轴标签
    ax.set_xlabel(t["time_label"], fontsize=12, fontproperties=zh_font if lang == "zh" else None)
    ax.set_ylabel(t["concentration_label"], fontsize=12, fontproperties=zh_font if lang == "zh" else None)
    title = "乳糖水解动力学" if lang == "zh" else "Lactose Hydrolysis Kinetics"
    ax.set_title(title, fontsize=14, fontproperties=zh_font if lang == "zh" else None)
    ax.grid(True, linestyle='--', alpha=0.7)
    ax.legend(loc='best', fontsize=10, prop=zh_font if lang == "zh" else None)
    ax.set_xlim([0, t_max])
    ax.set_ylim([0, L0 * 1.1])
    for spine in ax.spines.values():
        spine.set_linewidth(2.5)
    st.pyplot(fig)

    # 关键指标 - 显示所有抑制类型和无抑制的结果
    if all_results:
        # 创建结果表格
        results_data = []
        
        # 添加无抑制结果
        conversion_no_inh = (1 - L_no_inh[-1] / L0) * 100
        results_data.append({
            "抑制类型" if lang == "zh" else "Inhibition Type": t["no_inhibition"],
            t["final_lactose"]: f"{L_no_inh[-1]:.1f} mM",
            t["final_galactose"]: f"{Gal_no_inh[-1]:.1f} mM",
            t["conversion_rate"]: f"{conversion_no_inh:.1f}%"
        })
        
        # 添加选中的抑制类型结果
        for itype in inhibition_types:
            # 将显示名称映射到内部标识符
            if itype == t["competitive"]:
                key = "competitive"
                label = "竞争性抑制" if lang == "zh" else "Competitive"
            elif itype == t["non_competitive"]:
                key = "non_competitive"
                label = "非竞争性抑制" if lang == "zh" else "Non-competitive"
            elif itype == t["uncompetitive"]:
                key = "uncompetitive"
                label = "反竞争性抑制" if lang == "zh" else "Uncompetitive"
            else:
                continue
                
            t_hour, L, Gal, rates = all_results[key]
            conversion = (1 - L[-1] / L0) * 100
            results_data.append({
                "抑制类型" if lang == "zh" else "Inhibition Type": label,
                t["final_lactose"]: f"{L[-1]:.1f} mM",
                t["final_galactose"]: f"{Gal[-1]:.1f} mM",
                t["conversion_rate"]: f"{conversion:.1f}%"
            })
        
        # 显示结果表格
        results_df = pd.DataFrame(results_data)
        st.table(results_df)

    # 数据下载 - 包含所有情况的数据
    if all_results:
        # 创建Excel文件
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows

        wb = Workbook()
        # 移除默认创建的工作表
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        # 为每种情况添加工作表
        for inhibition_type, (t_hour, L, Gal, rates) in all_results.items():
            # 根据抑制类型确定工作表名称
            if inhibition_type == "no_inhibition":
                sheet_name = t["no_inhibition"]
            elif inhibition_type == "competitive":
                sheet_name = "竞争性抑制" if lang == "zh" else "Competitive"
            elif inhibition_type == "non_competitive":
                sheet_name = "非竞争性抑制" if lang == "zh" else "Non-competitive"
            elif inhibition_type == "uncompetitive":
                sheet_name = "反竞争性抑制" if lang == "zh" else "Uncompetitive"
            else:
                sheet_name = inhibition_type

            # 截断工作表名称（Excel限制31字符）
            sheet_name = sheet_name[:30]

            ws = wb.create_sheet(title=sheet_name)

            # 创建DataFrame
            if lang == "zh":
                df = pd.DataFrame({
                    '时间 (小时)': t_hour,
                    '乳糖浓度 (mM)': L,
                    '半乳糖浓度 (mM)': Gal,
                    '反应速率 (mM/小时)': rates
                })
            else:
                df = pd.DataFrame({
                    'Time (hours)': t_hour,
                    'Lactose (mM)': L,
                    'Galactose (mM)': Gal,
                    'Reaction Rate (mM/hour)': rates
                })

            # 将数据写入工作表
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)

        # 保存Excel文件
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        # 提供下载按钮
        st.download_button(
            label=t["download_data"],
            data=excel_buffer,
            file_name='lactose_hydrolysis_data.xlsx',
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    # 反应速率分析图 - 始终显示无抑制情况
    st.subheader(t["rate_analysis"])
    if all_results:  # 只要有无抑制结果就执行
        # 使用无抑制结果
        t_hour_no_inh, L_no_inh, Gal_no_inh, rates_no_inh = all_results["no_inhibition"]
        
        fig2, ax2 = plt.subplots(figsize=(10, 6))
        ax2.plot(L_no_inh, rates_no_inh, 'b--', linewidth=2.5,
                 label=t["no_inhibition"])
        
        # 找到最大反应速率及其发生时间（无抑制）
        max_rate_idx_no_inh = np.argmax(rates_no_inh)
        max_rate_no_inh = rates_no_inh[max_rate_idx_no_inh]
        max_rate_time_no_inh = t_hour_no_inh[max_rate_idx_no_inh]
        
        # 标注最大速率（无抑制）- 向上标注
        ax2.annotate(f'{t["no_inhibition"]} 最大速率: {max_rate_no_inh:.2f} mM/h' if lang == "zh" else f'{t["no_inhibition"]} Max rate: {max_rate_no_inh:.2f} mM/h',
                     xy=(L_no_inh[max_rate_idx_no_inh], max_rate_no_inh),
                     xytext=(L_no_inh[max_rate_idx_no_inh] + 0.05 * L0, max_rate_no_inh * 1.1),
                     arrowprops=dict(arrowstyle='->', color='blue'),
                     fontsize=10, fontproperties=zh_font if lang == "zh" else None)
        
        # 如果选择了抑制类型，添加第一个抑制类型的结果
        if inhibition_types:
            # 获取第一个抑制类型
            first_itype = inhibition_types[0]
            if first_itype == t["competitive"]:
                key = "competitive"
            elif first_itype == t["non_competitive"]:
                key = "non_competitive"
            elif first_itype == t["uncompetitive"]:
                key = "uncompetitive"
            else:
                key = "no_inhibition"
                
            if key in all_results:
                t_hour, L, Gal, rates = all_results[key]
                ax2.plot(L, rates, color=colors[key], linewidth=2.5,
                         label=first_itype)
                
                # 找到最大反应速率及其发生时间（抑制类型）
                max_rate_idx = np.argmax(rates)
                max_rate = rates[max_rate_idx]
                max_rate_time = t_hour[max_rate_idx]
                
                # 标注最大速率 - 向下标注
                annotation_text = f'{first_itype} 最大速率: {max_rate:.2f} mM/h' if lang == "zh" else f'{first_itype} Max rate: {max_rate:.2f} mM/h'
                ax2.annotate(annotation_text,
                             xy=(L[max_rate_idx], max_rate),
                             xytext=(L[max_rate_idx] + 0.05 * L0, max_rate * 0.7),
                             arrowprops=dict(arrowstyle='->', color='red'),
                             fontsize=10, fontproperties=zh_font if lang == "zh" else None)
                
                # 显示最大速率信息
                st.markdown(f"**{t['no_inhibition']}**: {t['max_rate'].format(max_rate_no_inh, max_rate_time_no_inh)}")
                st.markdown(f"**{first_itype}**: {t['max_rate'].format(max_rate, max_rate_time)}")
        
        # 如果没有选择抑制类型，显示无抑制的最大速率信息
        if not inhibition_types:
            st.markdown(f"**{t['no_inhibition']}**: {t['max_rate'].format(max_rate_no_inh, max_rate_time_no_inh)}")

        # 设置图表属性
        ax2.set_xlabel(t["substrate_label"], fontsize=12, fontproperties=zh_font if lang == "zh" else None)
        ax2.set_ylabel(t["rate_label"], fontsize=12, fontproperties=zh_font if lang == "zh" else None)
        title = f"反应速率 vs. 底物浓度" if lang == "zh" else "Reaction Rate vs. Substrate Concentration"
        ax2.set_title(title, fontsize=14, fontproperties=zh_font if lang == "zh" else None)
        ax2.grid(True, linestyle='--', alpha=0.7)
        ax2.legend(loc='best', prop=zh_font if lang == "zh" else None)
        ax2.set_xlim([0, L0])
        
        # 设置Y轴范围
        y_max = max(rates_no_inh) * 1.2
        if inhibition_types and key in all_results:
            y_max = max(y_max, max(rates) * 1.2)
        ax2.set_ylim([0, y_max])

        for spine in ax2.spines.values():
            spine.set_linewidth(2.5)
        st.pyplot(fig2)

    # Lineweaver-Burk 图表 - 始终显示
    st.subheader(t["lb_chart"])
    if all_results:  # 只要有无抑制结果就执行
        # 如果没有选择抑制类型，只显示无抑制线
        if not inhibition_types:
            # 只显示无抑制线
            Gal_fixed = st.slider(t["fixed_galactose"], 0.0, 200.0, 100.0)
            S_range = np.linspace(1, 500, 20)
            v_no_inh = Vmax * S_range / (Km + S_range)
            inv_S = 1 / S_range
            inv_v_no_inh = 1 / v_no_inh

            fig_lb, ax_lb = plt.subplots(figsize=(10, 6))
            p_no_inh = np.polyfit(inv_S, inv_v_no_inh, 1)
            x_fit_no_inh = np.linspace(-0.05, max(inv_S), 100)
            y_fit_no_inh = np.polyval(p_no_inh, x_fit_no_inh)
            ax_lb.plot(x_fit_no_inh, y_fit_no_inh, color='#4E6691', linewidth=2.5,
                       label="无抑制剂" if lang == "zh" else "No Inhibitor")

            # 计算截距
            y_intercept_no_inh = p_no_inh[1]
            x_intercept_no_inh = -p_no_inh[1] / p_no_inh[0]

            # 绘制截距点
            ax_lb.plot(0, y_intercept_no_inh, 'go', markersize=8, label="截距点" if lang == "zh" else "Intercepts")
            ax_lb.plot(x_intercept_no_inh, 0, 'b*', markersize=10)

            # 标注y轴截距（1/Vmax）
            ax_lb.annotate(r'$\frac{1}{V_{max}}$',
                           xy=(0, y_intercept_no_inh),
                           xytext=(0.01, y_intercept_no_inh - 1),
                           arrowprops=dict(arrowstyle='->', color='green'),
                           fontsize=12, color='green',
                           fontproperties=zh_font if lang == "zh" else None)

            # 标注x轴截距（-1/Km）
            ax_lb.annotate(r'$-\frac{1}{K_m}$',
                           xy=(x_intercept_no_inh, 0),
                           xytext=(x_intercept_no_inh, -1.5),
                           arrowprops=dict(arrowstyle='->', color='blue'),
                           fontsize=12, color='blue',
                           fontproperties=zh_font if lang == "zh" else None)

            # 设置坐标轴范围
            ax_lb.set_xlim(min(x_fit_no_inh)*1.1, max(x_fit_no_inh)*1.1)
            ax_lb.set_ylim(0, max(y_fit_no_inh)*1.1)
            
            # 设置标题和标签
            ax_lb.set_xlabel("1 / [S] (1/mM)", fontsize=12, fontproperties=zh_font if lang == "zh" else None)
            ax_lb.set_ylabel("1 / v (hour/mM)", fontsize=12, fontproperties=zh_font if lang == "zh" else None)
            title = "Lineweaver-Burk (无抑制)" if lang == "zh" else "Lineweaver-Burk (No Inhibition)"
            ax_lb.set_title(title, fontsize=14, fontproperties=zh_font if lang == "zh" else None)
            ax_lb.legend(loc='best', prop=zh_font if lang == "zh" else None)
            ax_lb.grid(True, linestyle='--', alpha=0.7)
            for spine in ax_lb.spines.values():
                spine.set_linewidth(2.5)
            st.pyplot(fig_lb)

            # 显示解释文本
            st.markdown("标准Michaelis-Menten动力学下的Lineweaver-Burk图" if lang == "zh" else "Lineweaver-Burk plot for standard Michaelis-Menten kinetics")
            
        else:
            # 如果有选择抑制类型，显示无抑制和第一个抑制类型的线
            # 获取第一个选择的抑制类型
            first_itype = inhibition_types[0]
            if first_itype == t["competitive"]:
                key = "competitive"
                display_key = t["competitive"]
            elif first_itype == t["non_competitive"]:
                key = "non_competitive"
                display_key = t["non_competitive"]
            elif first_itype == t["uncompetitive"]:
                key = "uncompetitive"
                display_key = t["uncompetitive"]
            else:
                key = "competitive"
                display_key = t["competitive"]

            Gal_fixed = st.slider(t["fixed_galactose"], 0.0, 200.0, 100.0)
            S_range = np.linspace(1, 500, 20)

            if key == "competitive":
                v_no_inh = Vmax * S_range / (Km + S_range)
                v_inh = Vmax * S_range / (Km * (1 + Gal_fixed / Ki) + S_range)
            elif key == "non_competitive":
                v_no_inh = Vmax * S_range / (Km + S_range)
                v_inh = Vmax * S_range / ((Km + S_range) * (1 + Gal_fixed / Ki))
            else:  # uncompetitive
                v_no_inh = Vmax * S_range / (Km + S_range)
                v_inh = Vmax * S_range / (Km + S_range * (1 + Gal_fixed / Ki))

            # 计算1/[S]和1/v
            inv_S = 1 / S_range
            inv_v_no_inh = 1 / v_no_inh
            inv_v_inh = 1 / v_inh

            fig_lb, ax_lb = plt.subplots(figsize=(10, 6))
            p_no_inh = np.polyfit(inv_S, inv_v_no_inh, 1)
            x_fit_no_inh = np.linspace(-0.05, max(inv_S), 100)
            y_fit_no_inh = np.polyval(p_no_inh, x_fit_no_inh)
            ax_lb.plot(x_fit_no_inh, y_fit_no_inh, color='#4E6691', linewidth=2.5,
                       label="无抑制剂" if lang == "zh" else "No Inhibitor")

            p_inh = np.polyfit(inv_S, inv_v_inh, 1)
            x_fit_inh = np.linspace(-0.05, max(inv_S), 100)
            y_fit_inh = np.polyval(p_inh, x_fit_inh)
            ax_lb.plot(x_fit_inh, y_fit_inh, color='#B8474D', linewidth=2.5,
                       label=f"{display_key}抑制" if lang == "zh" else f"{display_key} Inhibition")

            # 设置坐标轴范围
            all_x = np.concatenate([x_fit_no_inh, x_fit_inh])
            all_y = np.concatenate([y_fit_no_inh, y_fit_inh])
            
            # 设置X轴范围（添加10%的边距）
            x_min = min(all_x) * 1.1 if min(all_x) < 0 else min(all_x) * 0.9
            x_max = max(all_x) * 1.1
            ax_lb.set_xlim(x_min, x_max)
            
            # 设置Y轴范围（从0开始，添加10%的上边距）
            y_max = max(all_y) * 1.1
            ax_lb.set_ylim(0, y_max)

            # 计算截距
            y_intercept_no_inh = p_no_inh[1]
            y_intercept_inh = p_inh[1]
            x_intercept_no_inh = -p_no_inh[1] / p_no_inh[0]
            x_intercept_inh = -p_inh[1] / p_inh[0]

            # 绘制截距点
            ax_lb.plot(0, y_intercept_no_inh, 'go', markersize=8, label="截距点" if lang == "zh" else "Intercepts")
            ax_lb.plot(0, y_intercept_inh, 'ro', markersize=8)
            ax_lb.plot(x_intercept_no_inh, 0, 'b*', markersize=10)
            ax_lb.plot(x_intercept_inh, 0, 'r*', markersize=10)

            # 统一标注格式（与竞争性抑制相同）
            # 标注y轴截距（1/Vmax）
            ax_lb.annotate(r'$\frac{1}{V_{max}}$',
                           xy=(0, y_intercept_no_inh),
                           xytext=(0.01, y_intercept_no_inh - 1),
                           arrowprops=dict(arrowstyle='->', color='green'),
                           fontsize=12, color='green',
                           fontproperties=zh_font if lang == "zh" else None)

            # 标注有抑制的y轴截距
            if key == "competitive":
                # 竞争性抑制：y轴截距不变
                ax_lb.annotate(r'$\frac{1}{V_{max}}$',
                               xy=(0, y_intercept_inh),
                               xytext=(0.01, y_intercept_inh + 0.5),
                               arrowprops=dict(arrowstyle='->', color='red'),
                               fontsize=12, color='red',
                               fontproperties=zh_font if lang == "zh" else None)
            else:
                # 非竞争性和反竞争性抑制：y轴截距改变
                ax_lb.annotate(r'$\frac{1}{V_{max}^{app}}$',
                               xy=(0, y_intercept_inh),
                               xytext=(0.01, y_intercept_inh + 0.5),
                               arrowprops=dict(arrowstyle='->', color='red'),
                               fontsize=12, color='red',
                               fontproperties=zh_font if lang == "zh" else None)

            # 标注x轴截距（-1/Km）
            ax_lb.annotate(r'$-\frac{1}{K_m}$',
                           xy=(x_intercept_no_inh, 0),
                           xytext=(x_intercept_no_inh, -1.5),
                           arrowprops=dict(arrowstyle='->', color='blue'),
                           fontsize=12, color='blue',
                           fontproperties=zh_font if lang == "zh" else None)

            # 标注有抑制的x轴截距
            if key == "non_competitive":
                # 非竞争性抑制：x轴截距不变
                ax_lb.annotate(r'$-\frac{1}{K_m}$',
                               xy=(x_intercept_inh, 0),
                               xytext=(x_intercept_inh, -1.5),
                               arrowprops=dict(arrowstyle='->', color='red'),
                               fontsize=12, color='red',
                               fontproperties=zh_font if lang == "zh" else None)
            else:
                # 竞争性和反竞争性抑制：x轴截距改变
                ax_lb.annotate(r'$-\frac{1}{K_m^{app}}$',
                               xy=(x_intercept_inh, 0),
                               xytext=(x_intercept_inh, -1.5),
                               arrowprops=dict(arrowstyle='->', color='red'),
                               fontsize=12, color='red',
                               fontproperties=zh_font if lang == "zh" else None)

            ax_lb.set_xlabel("1 / [S] (1/mM)", fontsize=12, fontproperties=zh_font if lang == "zh" else None)
            ax_lb.set_ylabel("1 / v (hour/mM)", fontsize=12, fontproperties=zh_font if lang == "zh" else None)
            title = f"Lineweaver-Burk ({display_key}抑制)" if lang == "zh" else f"Lineweaver-Burk ({display_key} Inhibition)"
            ax_lb.set_title(title, fontsize=14, fontproperties=zh_font if lang == "zh" else None)
            ax_lb.legend(loc='best', prop=zh_font if lang == "zh" else None)
            ax_lb.grid(True, linestyle='--', alpha=0.7)
            for spine in ax_lb.spines.values():
                spine.set_linewidth(2.5)
            st.pyplot(fig_lb)

            # 显示解释文本
            st.markdown(t["lb_explanation"][key])

    # 练习题
    with st.expander(t["exercises"]):
        st.markdown(t["exercise_content"])

except Exception as e:
    st.error(t["error"].format(str(e)))
    st.stop()

st.caption(t["copyright"])
